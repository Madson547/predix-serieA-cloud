# ==========================================================
# Predix Série A — importar_qualitativos.py
# Lê a planilha predix_dados_qualitativos.xlsx (aba "Dados
# Qualitativos") e grava/atualiza os registros na tabela
# ajustes_qualitativos do Supabase (Pipeline A).
#
# Uso:
#   python importar_qualitativos.py
#   python importar_qualitativos.py "C:\caminho\outro_arquivo.xlsx"
# ==========================================================

import sys
from datetime import datetime, date
import openpyxl

from database import supabase

ARQUIVO_PADRAO = "predix_dados_qualitativos.xlsx"
ABA = "Dados Qualitativos"

# Chave da linha de exemplo que vem pronta na planilha — só usada pra
# avisar o usuário, não pra bloquear a importação.
LINHA_EXEMPLO = ("2026-08-02", "Flamengo", "Palmeiras")


def _to_str(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    texto = str(v).strip()
    return texto or None


def _to_float(v, default=1.0):
    try:
        if v is None or str(v).strip() == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _to_bool_sim_nao(v):
    if v is None:
        return False
    return str(v).strip().lower() in ("sim", "s", "true", "1")


def ler_planilha(caminho: str) -> list:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if ABA not in wb.sheetnames:
        raise ValueError(f"Aba '{ABA}' não encontrada em {caminho}")
    ws = wb[ABA]

    cabecalho = [c.value for c in ws[1]]
    registros = []
    avisou_exemplo = False

    for row in ws.iter_rows(min_row=2, values_only=True):
        linha = dict(zip(cabecalho, row))

        data_str = _to_str(linha.get("data"))
        time_casa = _to_str(linha.get("time_casa"))
        time_fora = _to_str(linha.get("time_fora"))

        if not data_str or not time_casa or not time_fora:
            continue  # linha vazia/incompleta — ignora

        if (data_str, time_casa, time_fora) == LINHA_EXEMPLO and not avisou_exemplo:
            print("[AVISO] A linha de exemplo (Flamengo x Palmeiras, 2026-08-02) ainda está "
                  "na planilha e será importada como se fosse real. Apague-a se não for um jogo de verdade.")
            avisou_exemplo = True

        registros.append({
            "data": data_str,
            "time_casa": time_casa,
            "time_fora": time_fora,
            "desfalques_casa": _to_str(linha.get("desfalques_casa")),
            "desfalques_fora": _to_str(linha.get("desfalques_fora")),
            "sequencia_casa": _to_str(linha.get("sequencia_casa")),
            "sequencia_fora": _to_str(linha.get("sequencia_fora")),
            "clima_vestiario_casa": _to_float(linha.get("clima_vestiario_casa"), 0.0),
            "clima_vestiario_fora": _to_float(linha.get("clima_vestiario_fora"), 0.0),
            "jogo_decisivo": _to_bool_sim_nao(linha.get("jogo_decisivo")),
            "ajuste_manual_casa": _to_float(linha.get("ajuste_manual_casa"), 1.0),
            "ajuste_manual_fora": _to_float(linha.get("ajuste_manual_fora"), 1.0),
            "observacoes": _to_str(linha.get("observacoes")),
            "resultado_real": _to_str(linha.get("resultado_real")),
            "placar_real": _to_str(linha.get("placar_real")),
            "data_atualizacao": datetime.now().isoformat(),
        })

    return registros


def gravar_supabase(registros: list):
    salvos = 0
    for row in registros:
        try:
            existe = supabase.table("ajustes_qualitativos").select("id") \
                .eq("data", row["data"]) \
                .eq("time_casa", row["time_casa"]) \
                .eq("time_fora", row["time_fora"]) \
                .execute()
            if existe.data:
                supabase.table("ajustes_qualitativos").update(row) \
                    .eq("data", row["data"]) \
                    .eq("time_casa", row["time_casa"]) \
                    .eq("time_fora", row["time_fora"]) \
                    .execute()
            else:
                supabase.table("ajustes_qualitativos").insert(row).execute()
            salvos += 1
        except Exception as e:
            print(f"[ERRO] {row['time_casa']} x {row['time_fora']} ({row['data']}): {e}")
    print(f"[IMPORTADOR] {salvos}/{len(registros)} registros salvos/atualizados.")


if __name__ == "__main__":
    caminho = sys.argv[1] if len(sys.argv) > 1 else ARQUIVO_PADRAO
    print(f"[IMPORTADOR] Lendo {caminho}...")
    registros = ler_planilha(caminho)
    print(f"[IMPORTADOR] {len(registros)} linhas válidas encontradas.")
    if registros:
        gravar_supabase(registros)
    else:
        print("[IMPORTADOR] Nada para importar.")
